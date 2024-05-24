from ast import Return
from cgitb import text
from faulthandler import disable
from selectors import EVENT_READ
from tkinter.tix import ComboBox
from turtle import left
import mysql.connector
# openpyxl son librerias para sacar datos de hojas de excel
from openpyxl import Workbook
from openpyxl import load_workbook
import openpyxl
# clase siguiente del tema de interface grafica
from tkinter import *
from tkinter import scrolledtext
from tkinter import messagebox
import tkinter as tk
from tkinter import ttk
from tkinter.ttk import Progressbar
from tkinter.ttk import *
from tkinter import *
from tkinter import messagebox, ttk
from datetime import datetime
from tkinter.ttk import Progressbar


#-----------------------------------------Funcion calcular edad--------------------------------------

def calc_edad():
  dbconexion = {
          'host':'localhost',#IP 169.254.215.159 si es un pc remoto
          'user':'root',
          'password':'',
          'db':'kits2'
    }

  conexion=mysql.connector.Connect(**dbconexion)
  cursor=conexion.cursor()
  sqlnum_registros="select count(Distancia) from `entregakits` where Distancia != ''"
  cursor.execute(sqlnum_registros)
  numRegistros=cursor.fetchall()
  i=0
  edad=0
  sqlnacimiento="SELECT * FROM entregakits"
  cursor.execute(sqlnacimiento)
  while i <= numRegistros[0][0]-1:

    fila=cursor.fetchone()
    fechanaci=fila[7]
    ceduladelencontrado=fila[3]
    #print(ceduladelencontrado)
    #cedula=str(ceduladelencontrado)
    if fechanaci is not None:
      #else:
      nacimiento=str(fechanaci)
      #print(nacimiento)
      formatofecha="%Y-%m-%d"
      d2 = datetime.strptime(nacimiento, formatofecha)
      d1 = datetime.strptime("2023-04-23", formatofecha)
      if d1.month > d2.month:
          edad= (d1.year - d2.year)
      if d1.month == d2.month:
          if d1.day >= d2.day:
              edad= (d1.year - d2.year)
          if d1.day < d2.day:
              edad=  (d1.year - d2.year)-1
      if d1.month < d2.month:
          edad= (d1.year - d2.year)-1 
    if fechanaci is None:
      edad="Sin Edad o problemas con fecha de nacimiento de CC:"+ceduladelencontrado
      #print (ceduladelencontrado)
      #print (i)
      messagebox.showinfo(message=edad, title=ceduladelencontrado)
    i+=1
    upgradeedad(ceduladelencontrado,edad)

#-----------------------------------------Funcion calcular categoria--------------------------------------
def calc_categoria():
  dbconexion = {
          'host':'localhost',#IP 169.254.215.159 si es un pc remoto
          'user':'root',
          'password':'',
          'db':'kits2'
    }

  conexion=mysql.connector.Connect(**dbconexion)
  cursor=conexion.cursor()
  sqlnum_registros="select count(Distancia) from `entregakits` where Distancia != ''"
  cursor.execute(sqlnum_registros)
  numRegistros=cursor.fetchall()
  i=0
  edad2=0
  sqlnacimiento="SELECT * FROM entregakits"
  cursor.execute(sqlnacimiento)
    
  while i <= numRegistros[0][0]-1:
    #bar=Progressbar(window, length=200, style='red.Horizontal.TProgressbar')
    #bar['value']=(i*numRegistros[0][0]-1)/100 #=73 % de avance de la barra de progreso
    #bar.place(x=170+i, y=390)
    
    
    fila=cursor.fetchone()
    Edad=fila[27]
    Dista=fila[20]
    Sexo=fila[6]
    ceduladelencontrado=fila[3]
    if Dista is not None:
      if Dista == "Vuelta a la Isla 32,5K":
        if Edad=="Sin Edad" : #or Edad=="No registr":
          categoria="Falta la edad, no permite calcular Categoria"
          messagebox.showinfo(message=categoria, title=ceduladelencontrado)
        else:
          edad2=int(Edad) 
          if Sexo == "Masculino":
              if edad2 <40:
                  categoria= "Mayores"
              if edad2 >=40 and edad2 <50:
                  categoria= "Master A"
              if edad2 >=50 and edad2 <60:
                  categoria= "Master B"
              if edad2 >=60:
                  categoria= "Master C"
          else:
              if edad2 < 40:
                      categoria="Abierta"
              if edad2 >=40 and edad2 <50:
                      categoria= "Plus"
              if edad2 >= 50:
                      categoria= "Master"
        
      else:
        categoria= "Unica"
            
    if Dista is None:
      mensajeDist="No registra Distancia"
      messagebox.showinfo(message=mensajeDist, title=ceduladelencontrado)
    
    i+=1
    upgradecategoria(ceduladelencontrado,categoria)
    
def calc_categoriaNew(ceduladelencontrado,Edad,Dista,Sexo):
    
  # para comentar varias filas de un solo guarapaso Contro K C y para descomentar control K U
  # dbconexion = {
  #         'host':'localhost',#IP 169.254.215.159 si es un pc remoto
  #         'user':'root',
  #         'password':'',
  #         'db':'kits2'
  #   }

  # conexion=mysql.connector.Connect(**dbconexion)
  # cursor=conexion.cursor()
  # sqlnum_registros="select count(Distancia) from `entregakits` where Distancia != ''"
  # cursor.execute(sqlnum_registros)
  # numRegistros=cursor.fetchall()
  # i=0
  # edad2=0
  # sqlnacimiento="SELECT * FROM entregakits"
  # cursor.execute(sqlnacimiento)
    
  # while i <= numRegistros[0][0]-1:
  #   #bar=Progressbar(window, length=200, style='red.Horizontal.TProgressbar')
  #   #bar['value']=(i*numRegistros[0][0]-1)/100 #=73 % de avance de la barra de progreso
  #   #bar.place(x=170+i, y=390)
    
    
   # fila=cursor.fetchone()
    
   # Edad=fila[27]
   # Dista=fila[20]
   # Sexo=fila[6]
   # ceduladelencontrado=fila[3]
    if Dista is not None:
      if Dista == "Vuelta a la Isla 32,5K":
        if Edad=="Sin Edad" : #or Edad=="No registr":
          categoria="Falta la edad, no permite calcular Categoria"
          messagebox.showinfo(message=categoria, title=ceduladelencontrado)
        else:
          edad2=int(Edad) 
          if Sexo == "Masculino":
              if edad2 <40:
                  categoria= "Mayores"
              if edad2 >=40 and edad2 <50:
                  categoria= "Master A"
              if edad2 >=50 and edad2 <60:
                  categoria= "Master B"
              if edad2 >=60:
                  categoria= "Master C"
          else:
              if edad2 < 40:
                      categoria="Abierta"
              if edad2 >=40 and edad2 <50:
                      categoria= "Plus"
              if edad2 >= 50:
                      categoria= "Master"
        
      else:
        categoria= "Unica"
            
    if Dista is None:
      mensajeDist="No registra Distancia"
      messagebox.showinfo(message=mensajeDist, title=ceduladelencontrado)
    
    #i+=1
    upgradecategoria(ceduladelencontrado,categoria)
    
#-----------------------------------------Funcion colocar los nombres a publicar o no autorizado--------------------------------------
    
def NombresPublicar():
  dbconexion = {
          'host':'localhost',#IP 169.254.215.159 si es un pc remoto
          'user':'root',
          'password':'',
          'db':'kits2'
    }

  conexion=mysql.connector.Connect(**dbconexion)
  cursor=conexion.cursor()
  sqlnum_registros="select count(Distancia) from `entregakits` where Distancia != ''"
  cursor.execute(sqlnum_registros)
  numRegistros=cursor.fetchall()
  i=0
  edad2=0
  sqlnacimiento="SELECT * FROM entregakits"
  cursor.execute(sqlnacimiento)
  while i <= numRegistros[0][0]-1:
    fila=cursor.fetchone()
    NombreP=fila[4]
    ApellidoP=fila[5]
    Autoriza=fila[22]
    ceduladelencontrado=fila[3] 
    if NombreP is not None:
      if Autoriza =="SI AUTORIZO":
        Publicarnombre= NombreP
        Publicarapellido=ApellidoP
      else: 
        Publicarnombre="Sin"
        Publicarapellido="Autorización"
    if NombreP is None:
      Publicarnombre="Sin Nombre"
      Publicarapellido="Sin Apellido"
      messagebox.showinfo(message=Publicarnombre+" "+Publicarapellido+" "+ceduladelencontrado, title=NombreP)
    
    i+=1
    upgradeNombrePublico(ceduladelencontrado,Publicarnombre,Publicarapellido)

def NombresPublicarNew(fila):
  # dbconexion = {
  #         'host':'localhost',#IP 169.254.215.159 si es un pc remoto
  #         'user':'root',
  #         'password':'',
  #         'db':'kits2'
  #   }

  # conexion=mysql.connector.Connect(**dbconexion)
  # cursor=conexion.cursor()
  # sqlnum_registros="select count(Distancia) from `entregakits` where Distancia != ''"
  # cursor.execute(sqlnum_registros)
  # numRegistros=cursor.fetchall()
  # i=0
  # edad2=0
  # sqlnacimiento="SELECT * FROM entregakits"
  # cursor.execute(sqlnacimiento)
  # while i <= numRegistros[0][0]-1:
  # fila=cursor.fetchone()
    NombreP=fila[4]
    ApellidoP=fila[5]
    print('nombre',NombreP,' ',ApellidoP)
    Autoriza=fila[22]
    ceduladelencontrado=fila[3] 
    if NombreP is not None:
      if Autoriza =="SI AUTORIZO":
        Publicarnombre= NombreP
        Publicarapellido=ApellidoP
      else: 
        Publicarnombre="Sin"
        Publicarapellido="Autorización"
    if NombreP is None:
      Publicarnombre="Sin Nombre"
      Publicarapellido="Sin Apellido"
      messagebox.showinfo(message=Publicarnombre+" "+Publicarapellido+" "+ceduladelencontrado, title=NombreP)
    
    #i+=1
    upgradeNombrePublico(ceduladelencontrado,Publicarnombre,Publicarapellido)


def upgradeedad(cedula1,edad1):
     dbconexion = {
          'host':'localhost',#IP 169.254.215.159 si es un pc remoto
          'user':'root',
          'password':'',
          'db':'kits2'
      }
     conexion=mysql.connector.Connect(**dbconexion)
     cursor=conexion.cursor()
     sql5="UPDATE entregakits SET Edad ='{}' WHERE Identificacion='{}'".format(edad1,cedula1 )
     cursor.execute(sql5)
     conexion.commit()

def upgradecategoria(cedula1,categoria1):
     dbconexion = {
          'host':'localhost',#IP 169.254.215.159 si es un pc remoto
          'user':'root',
          'password':'',
          'db':'kits2'
      }
     conexion=mysql.connector.Connect(**dbconexion)
     cursor=conexion.cursor()
     sql5="UPDATE entregakits SET Categ ='{}' WHERE Identificacion='{}'".format(categoria1,cedula1 )
     cursor.execute(sql5)
     conexion.commit()

def upgradeNombrePublico(cedula1,NombrePub1,ApellidoPub1):
     dbconexion = {
          'host':'localhost',#IP 169.254.215.159 si es un pc remoto
          'user':'root',
          'password':'',
          'db':'kits2'
      }
     conexion=mysql.connector.Connect(**dbconexion)
     cursor=conexion.cursor()
     sql5="UPDATE entregakits SET NombrePub ='{}', ApellidoPub='{}' WHERE Identificacion='{}'".format(NombrePub1,ApellidoPub1,cedula1 )
     cursor.execute(sql5)
     conexion.commit()



def ingresarBd():

 dbconexion = {
          'host':'localhost',#IP si es un pc remoto
          'user':'root',
          'password':'',
          'db':'kits2'
 }

 
 conexion=mysql.connector.Connect(**dbconexion)
 cursor=conexion.cursor()
 sql = "SELECT * FROM entregakits"
 cursor.execute(sql)
 resultados=cursor.fetchall()
 conexion.commit()
 
 #Ruta del archivo que se debe cargar a la base de datos
 #libro=load_workbook('C:/CarpetaCompartida/atletas.xlsx')
 libro=load_workbook('D:/Users/Edgar/San_Andres/2024/atletas.xlsx')
 
 hoja1=libro['Respuestas de formulario 1']
 c=hoja1['E'] # toma todos los datos de la columna E para sacar cuantas filas tiene 
 nfilas=len(c) # se obtiene el numero de filas de la columna E incluye la de los titulos
 nCol=30
 ultimaCelda='AD'+str(nfilas) #convierte a string el texto AE + ultima fila del acrivho de excel ej AE620
 print(ultimaCelda)
 rangoDatos=hoja1['A2':ultimaCelda] # crea el rango desde A2 hasta AE ultima fila de datos, ojo que el rango cuenta desde fila 0 col 0 

 # Abre la conexion antes del for
 for i in range(0,nfilas-1,1):
   conexion=mysql.connector.Connect(**dbconexion)
   cursor=conexion.cursor()
   # Para ingresar los datos de excel en la bd, se debe pasar a string, para ello en valores se indica la posicion con {} y en format se coloca la variable. 
   sql2 =("INSERT INTO entregakits VALUES ('{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}')".format('',rangoDatos[i][0].value, rangoDatos[i][1].value, rangoDatos[i][2].value, rangoDatos[i][3].value, rangoDatos[i][4].value, rangoDatos[i][5].value, rangoDatos[i][6].value, rangoDatos[i][7].value, rangoDatos[i][8].value, rangoDatos[i][9].value, rangoDatos[i][10].value, rangoDatos[i][11].value, rangoDatos[i][12].value, rangoDatos[i][13].value, rangoDatos[i][14].value, rangoDatos[i][15].value, rangoDatos[i][16].value, rangoDatos[i][17].value, rangoDatos[i][18].value, rangoDatos[i][19].value, rangoDatos[i][20].value, rangoDatos[i][21].value, rangoDatos[i][22].value, rangoDatos[i][23].value, rangoDatos[i][24].value, rangoDatos[i][25].value, rangoDatos[i][26].value, rangoDatos[i][27].value, rangoDatos[i][28].value, rangoDatos[i][29].value, '','Ninguno','0000',''))
   #sql2 ="INSERT INTO entregakits (FechaIns, Identificacion) VALUES ('"+rangoDatos[0][0].value+"','"+rangoDatos[0][1].value+ "')"
   cursor.execute(sql2)
   conexion.commit()
 print('Datos subidos exitosamente a la base de datos')
 #calc_edad()
 #print('calculó la edad')
 #calc_categoria()
 #print ('calculó categoria')
 #NombresPublicar()
 #print('calculó nombre publicar')
 
#  ********************************************************************* 
 
def ingresasrnewregistros():
  messagebox.showinfo(message="No olvide adicionar un fila al inicio en blanco y cambiar, en el Excel, el formato de la columna de cedulas a TEXTO. Y cambiar el nombre de la hoja a Respuestas de formulario 1. Si no lo ha hecho debe hacer el cambio a texto en el excel y luego regresar y dar aceptar", title="Advertencia")  
  dbconexion = {
          'host':'localhost',#IP 169.254.215.159 si es un pc remoto
          'user':'root',
          'password':'',
          'db':'kits2'
  }
  conexion=mysql.connector.Connect(**dbconexion)
  cursor2=conexion.cursor()
  sqlnum_registros="select count(Distancia) from `entregakits` where Distancia != ''"
  cursor2.execute(sqlnum_registros)
  numRegistros=cursor2.fetchall()#numero de registros de base datos PC
      #libro=load_workbook('C:/CarpetaCompartida/atletas.xlsx')
  libro=load_workbook('D:/Users/Edgar/San_Andres/2024/atletas.xlsx')
  
  hoja1=libro['Respuestas de formulario 1']
  c=hoja1['E'] # toma todos los datos de la columna E para sacar cuantas filas tiene 
 
  nfilas=len(c)-1 # se obtiene el numero de filas de la columna E incluye la de los titulos de la tabla atletas.xls nueva o descargada
  print ('n filas -1',nfilas)
  
 
 # ******************************************************** Codigo a midificar**********************
  
  #comparar nuemro de registro de las dos tablas y obtener ese dato
  nfilas_de_BD=numRegistros[0][0]# toco sumar 1 x q en mysql aparece que hay x filas, pero al exportar la tabla baja x+1 filas
  print ('nuemor filas de bd',nfilas_de_BD)
  CantidadRegistrosNuevos=nfilas - nfilas_de_BD

  print ('cantidad de registros nuevo',CantidadRegistrosNuevos)
  if CantidadRegistrosNuevos > 0:
    ConteoNuevosRegistros=CantidadRegistrosNuevos
    #nCol=30
    ultimaCelda='AD'+str(nfilas+1) #convierte a string el texto AE + ultima fila del acrivho de excel ej AE620
    rangoDatos=hoja1['A2':ultimaCelda] # crea el rango desde A2 hasta AE ultima fila de datos, ojo que el rango cuenta desde fila 0 col 0 
     
    for i in range(nfilas_de_BD,nfilas,1):
     print ('i=',i)
     print ('nfilas',nfilas)
     # Para ingresar los datos de excel en la bd, se debe pasar a string, para ello en valores se indica la posicion con {} y en format se coloca la variable. 
     #cursor=conexion.cursor()
     conexion=mysql.connector.Connect(**dbconexion)
     cursor=conexion.cursor()
     sql2 =("INSERT INTO entregakits VALUES ('{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}')".format('',rangoDatos[i][0].value, rangoDatos[i][1].value, rangoDatos[i][2].value, rangoDatos[i][3].value, rangoDatos[i][4].value, rangoDatos[i][5].value, rangoDatos[i][6].value, rangoDatos[i][7].value, rangoDatos[i][8].value, rangoDatos[i][9].value, rangoDatos[i][10].value, rangoDatos[i][11].value, rangoDatos[i][12].value, rangoDatos[i][13].value, rangoDatos[i][14].value, rangoDatos[i][15].value, rangoDatos[i][16].value, rangoDatos[i][17].value, rangoDatos[i][18].value, rangoDatos[i][19].value, rangoDatos[i][20].value, rangoDatos[i][21].value, rangoDatos[i][22].value, rangoDatos[i][23].value, rangoDatos[i][24].value, rangoDatos[i][25].value, rangoDatos[i][26].value, rangoDatos[i][27].value, rangoDatos[i][28].value, rangoDatos[i][29].value, '','Ninguno','0000',''))
     cursor.execute(sql2)
     #sql2 = """
     #        INSERT INTO entregakits
     #        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
     #"""
     #valores = ('', rangoDatos[i][0].value, rangoDatos[i][1].value, rangoDatos[i][2].value, rangoDatos[i][3].value, rangoDatos[i][4].value, rangoDatos[i][5].value, rangoDatos[i][6].value, rangoDatos[i][7].value, rangoDatos[i][8].value, rangoDatos[i][9].value, rangoDatos[i][10].value, rangoDatos[i][11].value, rangoDatos[i][12].value, rangoDatos[i][13].value, rangoDatos[i][14].value, rangoDatos[i][15].value, rangoDatos[i][16].value, rangoDatos[i][17].value, rangoDatos[i][18].value, rangoDatos[i][19].value, rangoDatos[i][20].value, rangoDatos[i][21].value, rangoDatos[i][22].value, rangoDatos[i][23].value, rangoDatos[i][24].value, rangoDatos[i][25].value, rangoDatos[i][26].value, rangoDatos[i][27].value, rangoDatos[i][28].value, rangoDatos[i][29].value, '', 'Ninguno', '0000', '')
     #cursor.execute(sql2, valores)
     conexion.commit()
     #conexion.close()
     print ('randgo col2',rangoDatos[i][2].value)
     print('Dato de Atleta subido exitosamente a la base de datos')
      
     #j=0
     edad=0
     #sqlnacimiento="SELECT * FROM entregakits"
     sqlnacimiento="SELECT * FROM entregakits ORDER BY ID DESC LIMIT %s"
     #cursor.execute(sqlnacimiento)
     cursor.execute(sqlnacimiento, (ConteoNuevosRegistros,))
     
     ConteoNuevosRegistros=ConteoNuevosRegistros-1
     
     #while j <= numRegistros[0][0]-1:
     #fila=cursor.fetchall()
     fila=cursor.fetchone()
     print ('la fila es:',fila)
     if fila:  
       fechanaci=fila[7]
       ceduladelencontrado=fila[3]
       Dista=fila[20]
       Sexo=fila[6]

       #print(ceduladelencontrado)
       #cedula=str(ceduladelencontrado)
       if fechanaci is not None:
            nacimiento=str(fechanaci)
            #print(nacimiento)
            formatofecha="%Y-%m-%d"
            d2 = datetime.strptime(nacimiento, formatofecha)
            d1 = datetime.strptime("2024-04-28", formatofecha)
            if d1.month > d2.month:
              edad= (d1.year - d2.year)
            if d1.month == d2.month:
                if d1.day >= d2.day:
                    edad= (d1.year - d2.year)
                if d1.day < d2.day:
                    edad=  (d1.year - d2.year)-1
            if d1.month < d2.month:
              edad= (d1.year - d2.year)-1 
        
       if fechanaci is None:
            edad="Sin Edad o problemas con fecha de nacimiento de CC:"+ceduladelencontrado
            #print (ceduladelencontrado)
            #print (i)
       #messagebox.showinfo(message=edad, title=ceduladelencontrado)
    
       
     upgradeedad(ceduladelencontrado,edad)
     calc_categoriaNew(ceduladelencontrado,edad,Dista,Sexo)
     NombresPublicarNew(fila)
  else:
    messagebox.showinfo(message="Ojo el numero de datos de la nueva tabla es menor que la tabla del pc NO HAY ATLETAS NUEVOS", title="Advertencia")  
  
  
  
       
  #calc_edad()
  #print('calculo la edad')
  #calc_categoria()
  #print ('calculo categoria')
  #NombresPublicar()
  #print('calculo nombre publicar')  
  print('Termino de correr el registro de nuevos excel')
  messagebox.showinfo(message="Termino de correr el registro de nuevos atletas", title="Advertencia")  
        
# fin de parte de programa que sube base de datos ************************************
#***********************************************************************************
#*******************************************************

def ventana4():
   global distCambiada
   def tomarSeleccion(distCambiada):
     distCambiada=distanNew.get()
     dbconexion = {
          'host':'localhost',#IP si es un pc remoto
          'user':'root',
          'password':'',
          'db':'kits2'
      }
     conexion=mysql.connector.Connect(**dbconexion)
     cursor=conexion.cursor()
     sql5="UPDATE entregakits SET Distancia ='{}' WHERE Identificacion='{}'".format(distCambiada,str(cedulasal) )
     cursor.execute(sql5)
     conexion.commit()
     window4.destroy()
     messagebox.showinfo(message="De aceptar y Espere 20 segundos hasta que aparezca el siguiente mensaje", title="cambio distancia")
     calc_categoria()
     messagebox.showinfo(message="Cambio Realizado Aceptar para continuar", title="cambio distancia")
     
     
     
   window4=Tk()
   window4.title("Cambio de Distancia")
   window4.geometry("300x200")
   tituloppal4=Label(window4,text="Cambio de Distancia", font=("Arial bold",20))
   tituloppal4.pack()
   distanNew= StringVar(window4)
   opciones=['Elija Distancia','Vuelta a la Isla 32,5K','21K (Categoría Única)','10K (Categoría Única)','5K (Categoría Única)']
   distanNew.set(opciones[0])
   opcion=OptionMenu(window4,distanNew,*opciones,command=tomarSeleccion)
   opcion.config(width=20)
   opcion.place(x=20, y=70)
   botoncerrar=Button(window4, text="No Cambiar Distancia", command=window4.destroy)
   botoncerrar.place(x=40, y=130)
   letrero_Camdist=Label(window4,text="Para verificar cambio, Cerrar, y buscar al mismo atleta", font=("Arial bold",10))
   letrero_Camdist.place(x=1, y=35)  

def ventana5():
   def eliminaratleta():
       dbconexion = {
           'host':'localhost',#IP si es un pc remoto
           'user':'root',
           'password':'',
           'db':'kits2'
       }
       conexion=mysql.connector.Connect(**dbconexion)
       cursor=conexion.cursor()
       sql6="UPDATE entregakits SET Retirado ='Retirado', Dorsal='', NameEntrega='', CelEntrega='' WHERE Identificacion='{}'".format(cedulasal )
       cursor.execute(sql6)
       conexion.commit()
       window5.destroy()
       
   window5=Tk()
   window5.title("Retiro de Atleta")
   window5.geometry("300x200")
   tituloppal5=Label(window5,text="Retiro de Atleta", font=("Arial bold",20))
   letrero_Camdist=Label(window5,text="¿Esta seguro de eliminar al Alleta?", font=("Arial bold",10))
   letrero_Camdist.place(x=1, y=35)  
   botoneliminar=Button(window5, text="Si, Eliminar", command=eliminaratleta)
   botoneliminar.place(x=40, y=80)
   botoncerrar2=Button(window5, text="No, Cerrar", command=window5.destroy)
   botoncerrar2.place(x=40, y=120)
   
def selection_changed():
    global numeroatleta
    numeroatleta=atletaSeleccion.get()
    for k in range(0,len(listaceduladelencontrado),1):
      if int(numeroatleta) == k+1:
       cedulasal=listaceduladelencontrado[k]
       dbconexion = {
          'host':'localhost',#IP si es un pc remoto
          'user':'root',
          'password':'',
          'db':'kits2'
        }  
       conexion=mysql.connector.Connect(**dbconexion)
       cursor=conexion.cursor()
       cedula.set(cedulasal)
       nombre_entrada.delete(0,"end")
       messagebox.showinfo(message="Ahora de clic en el Boton Buscar y se monstrarán los datos del atleta", title="Accón a Realizar")
       
# &&&&&&&&&&&&&&&&&&&&&&&&&&&& funcion busqueda &&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&

def busqueda():
  
 AtletaSeleccion_entrada.delete(0,"end") 
 global bandera
 bandera=1
 #Asignacion a una variable de  datos ingresados 
 global cedulasal
 cedulasal = str(cedula.get())
 nombresal=nombre.get()
 dbconexion = {
          'host':'localhost',#IP si es un pc remoto
          'user':'root',
          'password':'',
          'db':'kits2'
 }

 conexion=mysql.connector.Connect(**dbconexion)
 cursor=conexion.cursor()
  
 # *********** if que revisa la cedula   
 if cedulasal != '': 
    #messagebox.showinfo(message="Asegúrese de no ingresar la cedula con espacios intermedios", title="Ingreso Cedula")
    resultado2=[]
    sql = "SELECT Identificacion FROM entregakits WHERE Identificacion='{}'".format(str(cedulasal))
    cursor.execute(sql)
    resultado2=cursor.fetchall() # en la variable resultado2 queda la cedual encerrada en () ej (9770380)
    if len(resultado2) != 0: # si no encuentra la celdual en la bd, entonces la lista resultados2 es igual [] y su len = 0, entonces la cedula no existe
      asignadordal(cedulasal) # Esto me muestra en ventana los datos del atleta que se busca con cedula
    else:
      messagebox.showwarning(message="No lo encontro intente busqueda por nombre ", title="Ingreso Cedula")
      cedula_entrada.delete(0,"end")
      conexion.commit()
        
 else:
   messagebox.showwarning(message="Al dar aceptar, lo direccionará a otra ventana\n\nDeberá tomar el numero de la izquierda\n\nde la persona buscada, luego \n\nvolver a la página principal e ingresar ese número, en la casilla\n\n'Atleta Seleccionado' parte inferior derecha", title="Ingreso Cedula")
   cedula_entrada.delete(0,"end")
   conexion.commit()
   
  # ********* Fin del if que revisa la cedula
  #*********¨inicio busqueda por nombre o apellido
 
 if  nombresal != '':
  sqlnum_registros="select count(Distancia) from `entregakits` where Distancia != ''"
  cursor.execute(sqlnum_registros)
  numRegistros=cursor.fetchall()
  i=0
  conteo=0
  sql4="SELECT * FROM entregakits"
  cursor.execute(sql4)
  global listaNombres
  listaNombres=[]
  global listaposicionatletasenlista
  listaposicionatletasenlista=[]
  global listaceduladelencontrado
  listaceduladelencontrado=[]
  conteonombres=0
  while i <= numRegistros[0][0]-1:
    fila=cursor.fetchone()
    Resultadonombre=fila[4]+" "+fila[5]# se sumó el apellido para que en variable Resultadonombre quedo nombre y apellido
    ceduladelencontrado=fila[3]
    p=1
    longitudtexto=len(Resultadonombre)
    for j in range(0, longitudtexto,1):
      letranombre=Resultadonombre[j]
      if Resultadonombre[j] != " " and p==1:
        p=0
        Ninicio=j
       
      if Resultadonombre[j] == " " or j == longitudtexto-1:
        p=1
        Nfinal=j
        if j == longitudtexto-1:
          Nfinal=j+1
        nombreExtraido=Resultadonombre[Ninicio:Nfinal].upper()
        if nombresal.upper() == nombreExtraido:
          conteonombres += 1
          listaNombres.append(Resultadonombre)
          listaposicionatletasenlista.append(conteonombres)
          listaceduladelencontrado.append(ceduladelencontrado)
          
          j=longitudtexto*2
          #imprimir en ventana emergente fila[5] o la variable nombre  y que permita seleccionar      
          # y luego, despues de seleccionar aparezca los datos completos de inscripcion y aparezcan casillas de adignar dorsal  
          if conteo == 0:
             window3=Tk()
             window3.title("Bienvenido al sistema de asignación de Dorsal")
             window3.geometry("1400x2400")
             tituloppal3=Label(window3,text="Lista de nombres que coinciden", font=("Arial bold",20))
             tituloppal3.pack()
 
          textoconcatenado=str(conteonombres)+' '+Resultadonombre + '  ' + str(ceduladelencontrado)
          letrero_nombre=Label(window3,text=textoconcatenado, font=("Arial bold",10))
          #print('nombre escogido', conteonombres,' ', Resultadonombre)
          conteo += 20
          if conteo <= 600:  # 20*30 atletas q caben en la pantalla= 600 pixeles
             letrero_nombre.place(x=22, y=conteo)
          else:
             letrero_nombre.place(x=500, y=conteo-580)
    #fin del for         
    i+=1  
       
 else:
   #else sin usar  
  
  conexion.commit()
  banderalimpiar=1
  return 
 #------------------Fin funcion busqueda----------------

def verificasubirBd():
  if seleccion1.get() == 1:
    ingresarBd()    

def limpiarinicio():
   global letrero_Fecha # se debe poner global ya que en asingnardorsal() borramos el pantallazo y para q lo haga debe ser global
   global letrero_FechaDato
   global letrero_Cedula
   global letrero_CedulaDato
   global letrero_Nombre
   global letrero_NombreDato
   global letrero_Pago
   global letrero_PagoDato
   global letrero_Distancia
   global letrero_DistanciaDato
   global letrero_Talla
   global letrero_TallaDato
   global letrero_Dorsal
   global letrero_DorsalDato
   global letrero_Entregadoa
   global letrero_EntregadoaDato
   global letrero_EntregadoaCel
   global letrero_EntregadoaCelDato
   global letrero_Observ
   global letrero_ObservDato
   global boton_cambidistancia
   global boton_retirar 
   #limpiar()
      
def limpiar():
    DorsalAsignado=dorsalAsignado.get()
    if  DorsalAsignado == "":
      if BanderaAsignar == 0:
        #showinfo
        messagebox.showerror(message="¡ATENCIÓN! ¡ATENCIÓN! ¡ATENCIÓN!\n\n\nNO ASIGNÓ NINGUN DORSAL. SI DESAR ASIGANR EL DORSAL DEBERÁ BUSCAR NUEVAMENTE AL ATLETA", title="ADVERTENCIA")
      else:
        a= 'No hace nada'
        
      letrero_Fecha.config(text='')
      letrero_FechaDato.config(text='')
      letrero_Cedula.config(text='')
      letrero_CedulaDato.config(text='')
      letrero_Nombre.config(text='')
      letrero_NombreDato.config(text='')
      letrero_Pago.config(text='')
      letrero_PagoDato.config(text='')
      letrero_Distancia.config(text='')
      letrero_DistanciaDato.config(text='')
      letrero_Talla.config(text='')
      letrero_TallaDato.config(text='')
      letrero_Dorsal.config(text='')
      letrero_DorsalDato.config(text='')
      letrero_Entregadoa.config(text='')
      letrero_EntregadoaDato.config(text='')
      letrero_EntregadoaCel.config(text='')
      letrero_EntregadoaCelDato.config(text='')
      letrero_Observ.config(text='')
      letrero_ObservDato.config(text='')
      boton_cambidistancia.destroy()
      boton_retirar.destroy()
      boton_nuevoatleta.destroy()
      cedula_entrada.delete(0,"end")
      nombre_entrada.delete(0,"end")
    else: 
      messagebox.showerror(message="DIGITÓ UN DORSAL PERO NO SE LO ASIGNÓ. DEBE DAR ACEPTAR, ASIGNAR EL DORSAL Y DAR CLIC EN BOTÓN ASIGNAR DORSAL", title="ADVERTENCIA")
    
def actualizarDorsal():
    global DorsalAsignado
    DorsalAsignado=dorsalAsignado.get()
    print('dorsal asignado en la funcion actualizrDorsal:',dorsalAsignado.get())
    global Entregadoa
    Entregadoa=entregadoaa.get() #persona diferente a atleta al que se le entrego dorsal
    global Celularr
    Celularr=celularr.get() # celular del que se le entrego dorsal diferente a atleta
    
    sqldorsalbuscar = "SELECT Dorsal, Nombre, Apellido, Identificacion FROM entregakits WHERE Dorsal ='{}'".format(DorsalAsignado)
    cursor.execute(sqldorsalbuscar)
    dorsalbuscado=cursor.fetchall()
    #print('dorsal buscado:', dorsalbuscado)
    
    if len(dorsalbuscado) ==  0:
      sql3="UPDATE entregakits SET Dorsal ='{}', NameEntrega='{}', CelEntrega='{}' WHERE Identificacion='{}'".format(DorsalAsignado, Entregadoa, Celularr,cedulasal)
      cursor.execute(sql3)
      conexion.commit()
      messagebox.showinfo(message=f"Dorsal Asignado es {DorsalAsignado}", title="Dorsal Asignado")
      cedula_entrada.delete(0,"end")
      nombre_entrada.delete(0,"end")
      DorsalAsignado_entrada.delete(0,"end")
      Entregadoaa_entrada.delete(0,"end")
      Celularr_entrada.delete(0,"end")
      global BanderaAsignar
      BanderaAsignar=1
      limpiar()
      BanderaAsignar=0
    else: 
      messagebox.showinfo(message="El dorsal ya fue asignado a {} {} {}".format(dorsalbuscado[0][1],dorsalbuscado[0][2],dorsalbuscado[0][3]), title="Accón a Realizar")
    estadoentregadorsales()
    
def asignadordal(cedulasal):
   dbconexion = {
          'host':'localhost',#IP si es un pc remoto
          'user':'root',
          'password':'',
          'db':'kits2'
   }

   conexion=mysql.connector.Connect(**dbconexion)
   cursor=conexion.cursor()
   sql = "SELECT FechaIns, Identificacion, Nombre, Apellido, DPago, Distancia, Talla, Dorsal, NameEntrega, CelEntrega, Observ FROM entregakits WHERE Identificacion='{}'".format(cedulasal)
   cursor.execute(sql)
   resultado=cursor.fetchone() 
   global letrero_Fecha # se debe poner global ya que en asingnardorsal() borramos el pantallazo y para q lo haga debe ser global
   global letrero_FechaDato
   global letrero_Cedula
   global letrero_CedulaDato
   global letrero_Nombre
   global letrero_NombreDato
   global letrero_Pago
   global letrero_PagoDato
   global letrero_Distancia
   global letrero_DistanciaDato
   global letrero_Talla
   global letrero_TallaDato
   global letrero_Dorsal
   global letrero_DorsalDato
   global letrero_Entregadoa
   global letrero_EntregadoaDato
   global letrero_EntregadoaCel
   global letrero_EntregadoaCelDato
   global letrero_Observ
   global letrero_ObservDato
   global boton_cambidistancia
   global boton_retirar 
   global boton_nuevoatleta
   global BanderaAsignar
   BanderaAsignar=0
      
   letrero_Fecha=Label(window,text="Fecha Inscripción: ", font=("Arial bold",10))
   letrero_FechaDato=Label(window,text=resultado[0], font=("Arial bold",10))
   letrero_Cedula=Label(window,text="Cedula: ", font=("Arial bold",13))
   letrero_CedulaDato=Label(window,text=resultado[1], font=("Arial bold",13))
   letrero_Nombre=Label(window,text="Nombre: ", font=("Arial bold",13))
   letrero_NombreDato=Label(window,text=resultado[2]+' '+resultado[3], font=("Arial bold",15))
   letrero_Pago=Label(window,text="Forma de Pago: ", font=("Arial bold",13))
   letrero_PagoDato=Label(window,text=resultado[4], font=("Arial bold",13))
   letrero_Distancia=Label(window,text="Distancia: ", font=("Arial bold",20))
   letrero_DistanciaDato=Label(window,text=resultado[5], font=("Arial bold",20))
   letrero_Talla=Label(window,text="Talla: ", font=("Arial bold",20))
   letrero_TallaDato=Label(window,text=resultado[6], font=("Arial bold",20))
   letrero_Dorsal=Label(window,text="Dorsal Entregado: ", font=("Arial bold",15))
   letrero_DorsalDato=Label(window,text=resultado[7], font=("Arial bold",17))
   letrero_Entregadoa=Label(window,text="Dorsal Entregado a: ", font=("Arial bold",10))
   letrero_EntregadoaDato=Label(window,text=resultado[8], font=("Arial bold",10))
   letrero_EntregadoaCel=Label(window,text="Cel: ", font=("Arial bold",10))
   letrero_EntregadoaCelDato=Label(window,text=resultado[9], font=("Arial bold",10))
   letrero_Observ=Label(window,text="Observacion: ", font=("Arial bold",10))
   letrero_ObservDato=Label(window,text=resultado[10], font=("Arial bold",10))
      
   letrero_Fecha.place(x=20, y=370)
   letrero_FechaDato.place(x=130, y=370)
   letrero_Cedula.place(x=300, y=370)
   letrero_CedulaDato.place(x=360, y=370)
   letrero_Nombre.place(x=450, y=370)
   letrero_NombreDato.place(x=540, y=370)
   letrero_Pago.place(x=20, y=390)
   letrero_PagoDato.place(x=170, y=390)
   letrero_Distancia.place(x=20, y=450)
   letrero_DistanciaDato.place(x=150, y=450)
   letrero_Talla.place(x=450, y=450)
   letrero_TallaDato.place(x=550, y=450)
   letrero_Dorsal.place(x=20, y=500)
   letrero_DorsalDato.place(x=200, y=500)
   letrero_Entregadoa.place(x=280, y=500)
   letrero_EntregadoaDato.place(x=400, y=500)
   letrero_EntregadoaCel.place(x=550, y=500)
   letrero_EntregadoaCelDato.place(x=600, y=500)
   letrero_Observ.place(x=20, y=420)
   letrero_ObservDato.place(x=120, y=420)
   if resultado[7]:
    messagebox.showinfo(message="Este atleta ya tiene asignado el dorsal: " + str(resultado[7]) + " y fue entregado a " + resultado[8] + " cel: " + resultado[9], title="Advertencia")
    BanderaAsignar=1   
   else:
      print() #no hace nada
   #boton de limpiar de datos de pantalla para ingresar nuevo atleta 
   boton_nuevoatleta=Button(window, text="BUSCAR A OTRO ATLETA",command=limpiar, width="30", height="2",font=("Arial bold",11))
   boton_nuevoatleta.place(x=990,y=290)
   
   
   #boton para cambiar distancia que lo redirecciona a otra ventana (window4) para cambiar distnacia 
   boton_cambidistancia=Button(window, text="Cambiar Distancia",command=ventana4, width="20", height="1")
   boton_cambidistancia.place(x=900,y=500)
   
   #boton para eliminar atlteta que realmente no lo elimina, coloca la palabra "retirado" en col "Retirado" 
   boton_retirar=Button(window, text="Retirar Atleta",command=ventana5, width="20", height="1")
   boton_retirar.place(x=900,y=450)
   
   #conexion.commit()
   return cedulasal

def buscardorsal ():
    def buscardorsal2():
        dorsalabuscarget=dorsalabuscar.get()
        sqldorsalbuscar2 = "SELECT Dorsal, Nombre, Apellido, Identificacion FROM entregakits WHERE Dorsal ='{}'".format(dorsalabuscarget)
        cursor.execute(sqldorsalbuscar2)
        dorsalbuscado=cursor.fetchall()
   
        if len(dorsalbuscado) ==  0:
           messagebox.showinfo(message="El dorsal aun no ha sido asignado", title="Accón a Realizar")
           window6.destroy()
        else:
           messagebox.showinfo(message="El dorsal fue asignado a {} {} {}".format(dorsalbuscado[0][1],dorsalbuscado[0][2],dorsalbuscado[0][3]), title="Dorsal Asignado")
           window6.destroy()
      
    window6=Tk()
    window6.title("Buscar Dorsal")
    window6.geometry("300x200+10+120")
    tituloppal6=Label(window6,text="Dorsal  a Consultar", font=("Arial bold",20))
    letrero_buscardorsal=Label(window6,text="Ingrese Dorsal ", font=("Arial bold",15))
    letrero_buscardorsal.place(x=25, y=30) 

    dorsalabuscar=StringVar(window6)
      
    Buscardorsal_entrada=Entry(window6,textvariable=dorsalabuscar,width="20", font=("Arial bold",15) )
    Buscardorsal_entrada.place(x=50, y=60)
    Buscardorsal_entrada.focus()
    
    boton_buscardorsal2=Button(window6, text="Buscar",command=buscardorsal2, width="20", height="1", font=("Arial bold",13) )
    boton_buscardorsal2.place(x=50,y=100)
        
def newinscripcion():
    def distanciatomada(distTomada2):
        global distTomada
        distTomada=distanNewN.get()

    def sexotomado(sexoTomado2):
        global sexoTomado
        sexoTomado=sexo.get()
        #print('sexo tomado dentro:',sexoTomado)
    
    def sangretomado(sangreTomado2):
        global sangreTomado
        sangreTomado=sangre.get()
            
    def autorizatomado(autoriza2):
        global autorizaTomado
        autorizaTomado=autorizaN.get()
        
    def inscribir():
        sqlinsc =("INSERT INTO entregakits (Identificacion, Nombre, Apellido, FPago, DPago, Distancia,  Sexo, Sangre, FNacimiento,  Talla, email, Cel, Direccion, Ciudad, Pais, Autoriza, Club, ContactoEmer, Dorsal ) VALUES ('{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}')".format(cedulaN.get(), NombreN.get(),ApellidoN.get(),FPago.get(), DPago.get(),distTomada,sexoTomado,sangreTomado,fechaNaci.get(),tallaN.get(),emailN.get(),celularN.get(),direccN.get(),ciudadN.get(),paisN.get(),autorizaTomado,clubN.get(),cemerganciaN.get(),dorsal.get()))
        cursor.execute(sqlinsc)
        conexion.commit()
        calc_edad()
        print('calculó la edad')
        calc_categoria()
        print ('calculó categoria')
        NombresPublicar()
        print('calculó nombre publicar')
        window7.destroy()

    window7=Tk()
    window7.title("Nueva Inscripción")
    window7.geometry("1200x800+10+10")
    tituloppal7=Label(window7, text="Ingrese Nueva Inscripción", font=("Arial bold",20))
    tituloppal7.pack()
            
    letrero_CedulaN=Label(window7,text="Cedula: ", font=("Arial bold",13))
    letrero_CedulaN.place(x=20, y=40)
    cedulaN=StringVar(window7)
    cedulaN_entrada=Entry(window7, textvariable=cedulaN,width="15", font=("Arial bold",13) )
    cedulaN_entrada.place(x=20, y=60)
   
    letrero_NombreN=Label(window7,text="Nombre: ", font=("Arial bold",13))
    letrero_NombreN.place(x=20, y=90)
    NombreN=StringVar(window7)
    NombreN_entrada=Entry(window7, textvariable=NombreN,width="15", font=("Arial bold",13) )
    NombreN_entrada.place(x=20, y=110)
   
    letrero_ApellidoN=Label(window7,text="Apellido: ", font=("Arial bold",13))
    letrero_ApellidoN.place(x=200, y=90)
    ApellidoN=StringVar(window7)
    ApellidoN_entrada=Entry(window7, textvariable=ApellidoN,width="15", font=("Arial bold",13) )
    ApellidoN_entrada.place(x=200, y=110)
      
    letrero_FPago=Label(window7,text="Forma de Pago (Nequi, Daviplata, Efecty, Consignacion, Efectivo): ", font=("Arial bold",13))
    letrero_FPago.place(x=20, y=140)
    FPago=StringVar(window7)
    FPago_entrada=Entry(window7, textvariable=FPago,width="15", font=("Arial bold",13) )
    FPago_entrada.place(x=20, y=160)
   
    letrero_DPago=Label(window7,text="Detalle Pago: ", font=("Arial bold",13))
    letrero_DPago.place(x=20, y=190)
    DPago=StringVar(window7)
    DPago_entrada=Entry(window7, textvariable=DPago,width="15", font=("Arial bold",13) )
    DPago_entrada.place(x=20, y=210)
   
    # Distancia
    distanNewN= StringVar(window7)
    opciones=['Elija Distancia','Vuelta a la Isla 32,5K','21K (Categoría Única)','10K (Categoría Única)','5K (Categoría Única)']
    distanNewN.set(opciones[0])
    opcion=OptionMenu(window7,distanNewN,*opciones,command=distanciatomada)
    opcion.config(width=20)
    opcion.place(x=20, y=240)
    
    sexo= StringVar(window7)
    opcionessex=['Elija Genero','Masculino','Femenino']
    sexo.set(opcionessex[0])
    opcionsex=OptionMenu(window7,sexo,*opcionessex,command=sexotomado)
    opcionsex.config(width=20)
    opcionsex.place(x=200, y=240)
    
    sangre= StringVar(window7)
    opcionessangre=['Elija tipo de Sangre','A+','A-','B+','B-','AB+','AB-','O+','O-']
    sangre.set(opcionessangre[0])
    opcionsangre=OptionMenu(window7,sangre,*opcionessangre,command=sangretomado)
    opcionsangre.config(width=20)
    opcionsangre.place(x=400, y=240)
    
    letrero_fecha=Label(window7,text="Fecha Nacimento (aaaa-mm-dd) ej 1987-06-14", font=("Arial bold",13))
    letrero_fecha.place(x=20, y=270)
    fechaNaci=StringVar(window7)
    fechaNaci_entrada=Entry(window7, textvariable=fechaNaci,width="15", font=("Arial bold",13) )
    fechaNaci_entrada.place(x=20, y=290)
   
    letrero_talla=Label(window7,text="Talla Camiseta (s, m, l, xl)", font=("Arial bold",13))
    letrero_talla.place(x=20, y=320)
    tallaN=StringVar(window7)
    tallaN_entrada=Entry(window7, textvariable=tallaN,width="15", font=("Arial bold",13) )
    tallaN_entrada.place(x=20, y=340)
   
    letrero_email=Label(window7,text="email", font=("Arial bold",13))
    letrero_email.place(x=20, y=370)
    emailN=StringVar(window7)
    emailN_entrada=Entry(window7, textvariable=emailN,width="15", font=("Arial bold",13) )
    emailN_entrada.place(x=20, y=390)
    
    letrero_celularN=Label(window7,text="Celular", font=("Arial bold",13))
    letrero_celularN.place(x=20, y=420)
    celularN=StringVar(window7)
    celularN_entrada=Entry(window7, textvariable=celularN,width="15", font=("Arial bold",13) )
    celularN_entrada.place(x=20, y=440)
    
    letrero_direccN=Label(window7,text="Dirección", font=("Arial bold",13))
    letrero_direccN.place(x=20, y=470)
    direccN=StringVar(window7)
    direccN_entrada=Entry(window7, textvariable=direccN,width="15", font=("Arial bold",13) )
    direccN_entrada.place(x=20, y=490)
    
    letrero_ciudadN=Label(window7,text="Ciudad", font=("Arial bold",13))
    letrero_ciudadN.place(x=20, y=520)
    ciudadN=StringVar(window7)
    ciudadN_entrada=Entry(window7, textvariable=ciudadN,width="15", font=("Arial bold",13) )
    ciudadN_entrada.place(x=20, y=540)
    
    letrero_paisN=Label(window7,text="Pais", font=("Arial bold",13))
    letrero_paisN.place(x=220, y=520)
    paisN=StringVar(window7)
    paisN_entrada=Entry(window7, textvariable=paisN,width="15", font=("Arial bold",13) )
    paisN_entrada.place(x=220, y=540)
    
    letrero_autorizaN=Label(window7,text="Autoriaza publicar nombre: ", font=("Arial bold",13))
    letrero_autorizaN.place(x=20, y=570)
    autorizaN= StringVar(window7)
    opcionesautoriza=['Elija Opción','SI AUTORIZO','NO AUTORIZO']
    autorizaN.set(opcionesautoriza[0])
    opcionsautoriza=OptionMenu(window7,autorizaN,*opcionesautoriza,command=autorizatomado)
    opcionsautoriza.config(width=20)
    opcionsautoriza.place(x=220, y=570)
    
    letrero_clubN=Label(window7,text="Club", font=("Arial bold",13))
    letrero_clubN.place(x=20, y=600)
    clubN=StringVar(window7)
    clubN_entrada=Entry(window7, textvariable=clubN,width="15", font=("Arial bold",13) )
    clubN_entrada.place(x=20, y=620)
   
    letrero_contactoEmergencia=Label(window7,text="Contacto Emergencia y celular", font=("Arial bold",13))
    letrero_contactoEmergencia.place(x=220, y=620)
    cemerganciaN=StringVar(window7)
    cemergenciaN_entrada=Entry(window7, textvariable=cemerganciaN,width="15", font=("Arial bold",13) )
    cemergenciaN_entrada.place(x=500, y=620)
    
    letrero_dorsal=Label(window7,text="Dorsal a Asignar", font=("Arial bold",13))
    letrero_dorsal.place(x=670, y=620)
    dorsal=StringVar(window7)
    dorsal_entrada=Entry(window7, textvariable=dorsal,width="15", font=("Arial bold",13) )
    dorsal_entrada.place(x=800, y=620)
    
    boton_inscribir=Button(window7, text="Enviar",command=inscribir, width="20", height="1", font=("Arial bold",13) )
    boton_inscribir.place(x=960,y=620)
    
def crearlistcrono():
    #libro2=load_workbook('C:/Users/Usuario/Desktop/SAI2024/cronolist.xlsx')
    #libro2=load_workbook('C:/CarpetaCompartida/cronolist.xlsx')
    messagebox.showinfo(message="Por favor verifique que el archivo cornolist.xlsx, no tenga ningun tipo de dato")
    #hoja1=libro2['Sheet']
    libro2=Workbook()
    hoja1=libro2.active
    
    sqlnum_registros="select count(Distancia) from `entregakits` where Distancia != ''"
    cursor.execute(sqlnum_registros)
    numRegistros=cursor.fetchall()
    i=0
    print('Entra a la funcion y cuenta datos existentes en tabla')
    print(numRegistros[0][0])
    sql4="SELECT * FROM entregakits"
    cursor.execute(sql4)
    while i <= numRegistros[0][0]-1: 
       fila=cursor.fetchone()
       if  fila[31] != 'None' : #posocion 31 es dorsal or fila[31] != '' :#or fila[29] != 'Retirado':
         if fila[31] != '':
            #if fila[34] == '' or fila[34]=='None': # posicion 34 es retirado
            #  textobasura='texto para no hacer nada'
            #  print ('entro a que es diferente')
            #  print (fila[34])
            #else: 
              nombreMasApellido=fila[29]+' '+fila[30]   
              #Resultadonombre=fila[26].split()
              Resultadonombre=nombreMasApellido.split()
              #Resultadonombrecompleto=fila[26]
              Resultadonombrecompleto=nombreMasApellido
              numerodenombres=len(Resultadonombre) # si es una cadena entonces bota numero de caracteres si es una lista bota el numero de elementos que es nuestro caso
              if numerodenombres < 5:
                if numerodenombres == 2:
                  #paser el nombre y apellido
                  dorsal=fila[31]
                  nombre=Resultadonombre[0]
                  apellido=Resultadonombre[1]
                  genero=fila[6]
                  edad=fila[27]
                  distancia=fila[20]
                  categoria=fila[28]
                  renglon=[Resultadonombrecompleto, dorsal,nombre,apellido,genero,edad,distancia,categoria]
                  hoja1.append(renglon)
                                    
                elif numerodenombres == 3 :
                  dorsal=fila[31]
                  nombre=Resultadonombre[0]
                  apellido=Resultadonombre[1]
                  genero=fila[6]
                  edad=fila[27]
                  distancia=fila[20]
                  categoria=fila[28]
                  renglon=[Resultadonombrecompleto,dorsal,nombre,apellido,genero,edad,distancia,categoria,Resultadonombrecompleto]
                  hoja1.append(renglon)
                 
                elif numerodenombres == 4:
                  dorsal=fila[31]
                  nombre=Resultadonombre[0]
                  apellido=Resultadonombre[2]
                  genero=fila[6]
                  edad=fila[27]
                  distancia=fila[20]
                  categoria=fila[28]
                  renglon=[Resultadonombrecompleto,dorsal,nombre,apellido,genero,edad,distancia,categoria]
                  hoja1.append(renglon)
                
                else:
                    Texto='solo poner el else sin hacer nada'  
                    
              else:  
                dorsal=fila[31]
                nombre=Resultadonombre[0]
                apellido=''
                genero=fila[6]
                edad=fila[27]
                distancia=fila[20]
                categoria=fila[28]
                renglon=[Resultadonombrecompleto,dorsal,nombre,apellido,genero,edad,distancia,categoria,Resultadonombrecompleto]
                hoja1.append(renglon)
                
         else:
               Texto='solo poner el else sin hacer nada'           
       else:
         Texto='solo poner el else sin hacer nada'  
       i += 1
       
    # *************************fin del while
    messagebox.showinfo(message="El Archivo ya fue creado")
    libro2.save('C:/Users/Usuario/Desktop/SAI2024/cronolist.xlsx')
    libro2.close()
    #libro2.save('C:/CarpetaCompartida/cronolist.xlsx')

    #c=hoja1['E'] 

def estadoentregadorsales():
    letreroDorEntre.config(text='')
    letreroDorxEntre.config(text='')
    dbconexion = {
          'host':'localhost',#169.254.215.159IP si es un pc remoto
          'user':'root',
          'password':'',
          'db':'kits2'
    }
    conexion=mysql.connector.Connect(**dbconexion)
    cursor=conexion.cursor()
    sqldorsalesentregados="select count(Dorsal) from `entregakits` where Dorsal>0"

    cursor.execute(sqldorsalesentregados)
    DorEntre=cursor.fetchall()

    letrero_dor_entreg=Label(window,text="Dorsales Entregados: ", font=("Arial bold",15))
    letrero_dor_entreg.place(x=22, y=160)
    letreroDorEntre=Label(window,text=DorEntre[0], font=("Arial bold",15))
    letreroDorEntre.place(x=230,y=160)
    conexion.commit()

    dbconexion = {
          'host':'localhost',#169.254.215.159 IP si es un pc remoto
          'user':'root',
          'password':'',
          'db':'kits2'
     }
    conexion2=mysql.connector.Connect(**dbconexion)
    cursor2=conexion2.cursor()
    sqldorsalesxentregar="select count(Dorsal) from `entregakits` where Dorsal=0"
    cursor2.execute(sqldorsalesxentregar)
    DorxEntregar=cursor2.fetchall()
    conexion.commit()

    letrero_dor_xentreg=Label(window,text="Dorsales por Entregar: ", font=("Arial bold",15))
    letrero_dor_xentreg.place(x=422, y=160)
    letreroDorxEntre=Label(window,text=DorxEntregar[0], font=("Arial bold",15))
    letreroDorxEntre.place(x=822,y=160)    
#************************ fin de funciones*********************************


#                                                                         #  
#                                                                         #
#                                                                         #
#                                                                         #
#                                                                         #
#                                                                         #
#                                                                         #
#                                                                         #
#*************************************************************************#


 
# *******************Inicio programa subir base datos ****************************** 


window=Tk()
window.title("Bienvenido al sistema de asignación de Dorsal")
window.geometry("1400x1400")
#window.config(background="#213241")
tituloppal=Label(window,text="Subir Base de Datos", font=("Arial bold",20))
tituloppal.pack()

tituloIngresoBd=Label(window,text="Si ya se subió la base de datos siga con Asignación de Dorsal ", font=("Arial bold",15))
tituloIngresoBd.pack()
seleccion1=IntVar()
seleccion1.set(2) # Permite que la opcion 2 siempre quede marcada automaticamente

radi1=Radiobutton(window,text='Ingresar Base Datos',variable=seleccion1, value=1,)
radi1.place(x=10, y=70)
radi2=Radiobutton(window,text='Ya esta ingresada la Base Datos',variable=seleccion1, value=2,)
radi2.place(x=10, y=100)

boton_ingresoBD=Button(window, text="Ingresar Base Datos",command=verificasubirBd, width="20", height="1")
boton_ingresoBD.place(x=150,y=70)

boton_ingresoBDnewRegistros=Button(window, text="Ingresar Nuevos a  BD",command=ingresasrnewregistros, width="20", height="1")
boton_ingresoBDnewRegistros.place(x=450,y=70)


# ************************ fin programa subir base de datos*********************************

# -------------------- inicio del programa de asignacion de dorsal --------------------------------------------------------------------------------

tituloIngresoDorsal=Label(window,text="------------------------------ASIGNACION DE DORSAL------------------------------", font=("Arial bold",20))
tituloIngresoDorsal.place(x=10, y=130)

# Definicion de variables glbales que se calculan en las def y q las necesitamos en todo el programa
cedulasal=''# definimos esta variable globar de string sin string para luego tomar el dato que es modificado en una def  
DorsalAsignado=''
Entregadoa=''
Celularr=''
atletaSeleccion='provisional'
bandera2=0
bandera=0
distCambiada=''
global banderalimpiar
banderalimpiar=0 
distTomada='xxxxx'
sexoTomado='Masculinoooo' 
posicion=''
limpiarinicio()


dbconexion = {
          'host':'localhost',#169.254.215.159IP si es un pc remoto
          'user':'root',
          'password':'',
          'db':'kits2'
 }
conexion=mysql.connector.Connect(**dbconexion)
cursor=conexion.cursor()
sqldorsalesentregados="select count(Dorsal) from `entregakits` where Dorsal>0"

cursor.execute(sqldorsalesentregados)
DorEntre=cursor.fetchall()

letrero_dor_entreg=Label(window,text="Dorsales Entregados: ", font=("Arial bold",15))
letrero_dor_entreg.place(x=22, y=160)

letreroDorEntre=Label(window,text=DorEntre[0], font=("Arial bold",15))
letreroDorEntre.place(x=230,y=160)
conexion.commit()

dbconexion = {
          'host':'localhost',#169.254.215.159IP si es un pc remoto
          'user':'root',
          'password':'',
          'db':'kits2'
 }
conexion2=mysql.connector.Connect(**dbconexion)
cursor2=conexion2.cursor()
sqldorsalesxentregar="select count(Dorsal) from `entregakits` where Dorsal=0"
cursor2.execute(sqldorsalesxentregar)
DorxEntregar=cursor2.fetchall()
conexion.commit()

letrero_dor_xentreg=Label(window,text="Dorsales por Entregar: ", font=("Arial bold",15))
letrero_dor_xentreg.place(x=422, y=160)
letreroDorxEntre=Label(window,text=DorxEntregar[0], font=("Arial bold",15))
letreroDorxEntre.place(x=822,y=160)

letrero_cedula=Label(window,text="Cédula(sin puntos ni espacios)", font=("Arial bold",15))
letrero_cedula.place(x=22, y=260)
letrero_nombre=Label(window,text=" Nombre ó Apellido", font=("Arial bold",15))
letrero_nombre.place(x=300, y=260)

cedula=StringVar()
nombre=StringVar()

cedula_entrada=Entry(textvariable=cedula,width="20", font=("Arial bold",15) )
cedula_entrada.place(x=22, y=290)
cedula_entrada.focus()

nombre_entrada=Entry(textvariable=nombre,width="20", font=("Arial bold",15))
nombre_entrada.place(x=300,y=290)

boton_buscar=Button(window, text="Buscar",command=busqueda, width="20", height="1", font=("Arial bold",15))
boton_buscar.place(x=990,y=290)

# ************************* Captura de dorsal y de a quien se lo entrego

letrero_DorsalAsignado=Label(text="Asignar Dorsal", font=("Arial bold",15))
letrero_DorsalAsignado.place(x=20, y=600)

dorsalAsignado=StringVar()
   
DorsalAsignado_entrada=Entry(textvariable=dorsalAsignado,width="10", font=("Arial bold",15) )
DorsalAsignado_entrada.place(x=20, y=640)
    
letrero_Entregadoaa=Label(window,text="Entregado a: ", font=("Arial bold",15))
letrero_Entregadoaa.place(x=250, y=600) 

entregadoaa=StringVar()
      
Entregadoaa_entrada=Entry(window,textvariable=entregadoaa,width="20", font=("Arial bold",15) )
Entregadoaa_entrada.place(x=250, y=640)
   
letrero_Celularr=Label(window,text="Celular: ", font=("Arial bold",15))
letrero_Celularr.place(x=550, y=600)

celularr=StringVar() #celular de quien recibe el kit si es diferente al corredor
   
Celularr_entrada=Entry(window,textvariable=celularr,width="20", font=("Arial bold",15) )
Celularr_entrada.place(x=550, y=640)

boton_actualizarDorsal=Button(window, text="Asignar Dorsal",command=actualizarDorsal, width="20", height="1", font=("Arial bold",13) )
#boton_actualizarDorsal.place(x=990,y=290)
boton_actualizarDorsal.place(x=800,y=640)



# *****************************este boton finaliza el programa
boton_finalizar=Button(window, text="Terminar",command=exit, width="20", height="1", font=("Arial bold",13) )
boton_finalizar.place(x=1150,y=640)

#***************************** fin de Captura de dorsal y de a quien se lo entrego

# *********************** inicio de la toma de numero del atleta buscado por nombre************

letrero_AtletaSeleccion=Label(window, text="Atleta Seleccionado ", font=("Arial bold",13))
letrero_AtletaSeleccion.place(x=950, y=550)
            
atletaSeleccion=StringVar()
            
AtletaSeleccion_entrada=Entry(window, textvariable=atletaSeleccion,width="10", font=("Arial bold",13) )
AtletaSeleccion_entrada .place(x=1100, y=550)
            
boton_atletaSeleccion=Button(window, text="Enviar",command=selection_changed, width="20", height="1", font=("Arial bold",13) )
boton_atletaSeleccion.place(x=950,y=580)

# *****************************   boton PARA BUSCAR DORSAL    **********************
boton_buscardorsal=Button(window, text="Buscar Dorsal",command=buscardorsal, width="20", height="1", font=("Arial bold",10) )
boton_buscardorsal.place(x=990,y=250)

# *****************************   boton para nuevo inscrito    **********************
boton_newinscrito=Button(window, text="Inscripción Nuevo Atleta",command=newinscripcion, width="20", height="1", font=("Arial bold",10) )
boton_newinscrito.place(x=990,y=200)

boton_crearListaCrono=Button(window, text="Crear Listado Crono",command=crearlistcrono, width="20", height="1", font=("Arial bold",10) )
boton_crearListaCrono.place(x=990,y=150)

window.mainloop()
#---------------------------------FIN CODIGO-------------------------------------------
